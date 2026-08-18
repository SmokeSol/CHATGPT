#!/usr/bin/env python3
from __future__ import annotations

import hashlib, itertools, json, math, re, tempfile, unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import numpy as np
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
A=ROOT/'data'/'goal100'/'agent_society_v2'
H=ROOT/'data'/'goal100'/'historical'
GEO=ROOT/'data'/'goal100'/'geometry_2026_certificate.json'
OUT=A/'populations'; OUT.mkdir(parents=True,exist_ok=True)
CERT=A/'population_certificate_v1.json'
HCP_URL='https://www.rgph2014.hcp.ma/file/190479/'
CORE=('RNI','PAM','PI','PJD','USFP','MP','UC','PPS')
PARTIES=(*CORE,'OTHER')
PRIOR_STATES=(*PARTIES,'ABSTAIN')
AGE=['18_24','25_34','35_44','45_59','60_PLUS']
SEX=['M','F']; URBAN=['URBAN','RURAL']
EDU=['NONE_OR_PRESCHOOL','PRIMARY','SECONDARY','HIGHER']
ACT=['ACTIVE_EMPLOYED','UNEMPLOYED','INACTIVE']
TARGET_N=128

ALIASES={
'rabat el mouhit':'rabat ocean','rabat al mouhit':'rabat ocean','rabat challah':'rabat chellah',
'fes janoubia':'fes sud','fes chamalia':'fes nord','fes shamalia':'fes nord',
'marrakech medina':'medina sidi youssef','marrakech gueliz ennakhil':'gueliz nakhil','marrakech gueliz nakhil':'gueliz nakhil',
'marrakech menara':'menara','moulay yaacoub':'moulay yaacoub','moulay yacoub':'moulay yaacoub',
'm diq fnideq':'m diq fnideq','taroudannt al janoubia':'taroudant sud','taroudannt chamalia':'taroudant nord'
}

def canon(o): return json.dumps(o,ensure_ascii=False,sort_keys=True,separators=(',',':'),allow_nan=False).encode()
def sha(b): return hashlib.sha256(b).hexdigest()
def read(p): return json.loads(p.read_text(encoding='utf-8'))
def norm(s):
    x=unicodedata.normalize('NFKD',str(s or '')); x=''.join(c for c in x if not unicodedata.combining(c))
    x=x.lower().replace('’',"'"); x=re.sub(r'[^a-z0-9]+',' ',x); x=re.sub(r'\s+',' ',x).strip(); return ALIASES.get(x,x)
def sim(a,b):
    if a==b:return 1.0
    sa,sb=set(a.split()),set(b.split()); j=len(sa&sb)/max(1,len(sa|sb)); seq=SequenceMatcher(None,a,b).ratio(); return .58*seq+.42*j

def load_hist(year):
    d=read(H/f'tafra_legislative_{year}_canonical.json')
    rows=[r for r in d['rows'] if str(r.get('list_type','')).lower() in {'locale','local'}]
    if len(rows)!=92: raise RuntimeError(f'historical {year}: expected 92 local rows, got {len(rows)}')
    return rows

def match_geometry_to_hist(geo_rows,hrows):
    pairs=[]
    for c in geo_rows:
        cn=norm(c['repo_name'])
        for h in hrows:
            hn=norm(h['constituency']); score=sim(cn,hn)
            try: score += .06 if int(h.get('seats'))==int(c['repo_seats']) else -.08
            except Exception: pass
            pairs.append((score,c['constituency_id'],str(h['id_constituency']),h))
    pairs.sort(key=lambda x:(-x[0],x[1],x[2])); uc=set();uh=set();out={};audit={}
    for score,cid,hid,h in pairs:
        if cid in uc or hid in uh or score<.43: continue
        uc.add(cid);uh.add(hid);out[cid]=h;audit[cid]={'id_constituency':hid,'score':score,'historical_name':h['constituency']}
    if len(out)!=92: raise RuntimeError(f'geometry/history mapping only {len(out)}/92')
    return out,audit

def bucket_shares(row):
    raw=row.get('votes',{}); total=sum(float(v or 0) for v in raw.values())
    if total<=0: raise RuntimeError('zero valid votes')
    d={p:float(raw.get(p,0) or 0)/total for p in CORE}; d['OTHER']=sum(float(v or 0) for k,v in raw.items() if k not in CORE)/total
    s=sum(d.values()); return {k:v/s for k,v in d.items()}

def political_margin(row):
    t=float(row.get('turnout_rate_reported'))
    if not 0<t<1: raise RuntimeError(f'invalid turnout {t} for {row.get("constituency")}')
    sh=bucket_shares(row); d={p:t*sh[p] for p in PARTIES}; d['ABSTAIN']=1-t
    s=sum(d.values()); return {k:v/s for k,v in d.items()}

def fetch_hcp():
    r=requests.get(HCP_URL,timeout=(20,150),headers={'User-Agent':'Atlas395-ASV2/1.0'});r.raise_for_status();b=r.content
    if not b.startswith(b'PK'): raise RuntimeError('HCP response not xlsx')
    return b

def province_key(s):
    x=norm(s); x=re.sub(r'^(province|prefecture)\s+','',x).strip(); return x

def hcp_tables(b):
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
        f.write(b);f.flush(); wb=load_workbook(f.name,read_only=True,data_only=True)
        tabs={}
        for sn in ['Indic.Ensemble','Indic.Urbain','Indic.Rural']:
            ws=wb[sn]; rows={}
            for row in ws.iter_rows(min_row=4,values_only=True):
                label=row[7] if len(row)>7 else None
                if label and str(label).lower().startswith(('province:','préfecture:','prefecture:')):
                    rows[province_key(label)]=row
            tabs[sn]=rows
    return tabs

def pct(row,col1):
    v=row[col1-1]
    if v in (None,''): return 0.0
    x=float(v); return max(0.0,x)/100.0

def num(row,col1):
    v=row[col1-1]; return 0.0 if v in (None,'') else float(v)

def adult_age_counts(row,pop_col,age_start_col):
    pop=num(row,pop_col)
    # 15-19 is age_start+3, 20-24 +4, ..., 75+ +15. Eligible 18-19 assumed 2/5 of 15-19 uniformly.
    q=[pct(row,age_start_col+i) for i in range(16)]
    d={
      '18_24':pop*(.4*q[3]+q[4]),
      '25_34':pop*(q[5]+q[6]),
      '35_44':pop*(q[7]+q[8]),
      '45_59':pop*(q[9]+q[10]+q[11]),
      '60_PLUS':pop*sum(q[12:16]),
    }
    return d

def normalize_dict(d):
    s=sum(max(0.0,float(v)) for v in d.values())
    if s<=0: raise RuntimeError(f'zero marginal {d}')
    return {k:max(0.0,float(v))/s for k,v in d.items()}

def demo_margin(province,tabs):
    key=province_key(province)
    ens=tabs['Indic.Ensemble'].get(key); urb=tabs['Indic.Urbain'].get(key); rur=tabs['Indic.Rural'].get(key)
    if ens is None: raise RuntimeError(f'HCP province missing: {province} key={key}')
    age=normalize_dict(adult_age_counts(ens,10,11))
    male=sum(adult_age_counts(ens,64,65).values()); female=sum(adult_age_counts(ens,118,119).values())
    sex=normalize_dict({'M':male,'F':female})
    def adult_total(row): return 0.0 if row is None else sum(adult_age_counts(row,10,11).values())
    urban=normalize_dict({'URBAN':adult_total(urb),'RURAL':adult_total(rur)})
    edu=normalize_dict({
      'NONE_OR_PRESCHOOL':pct(ens,41)+pct(ens,42),
      'PRIMARY':pct(ens,43),
      'SECONDARY':pct(ens,44)+pct(ens,45),
      'HIGHER':pct(ens,46),
    })
    active=pct(ens,52); inactive=pct(ens,53); u=max(0.0,min(1.0,pct(ens,55)))
    activity=normalize_dict({'ACTIVE_EMPLOYED':active*(1-u),'UNEMPLOYED':active*u,'INACTIVE':inactive})
    return {'age_band':age,'sex':sex,'urban_rural':urban,'education_band':edu,'activity_status':activity}, {'hcp_key':key,'urban_row_present':urb is not None,'rural_row_present':rur is not None,'eligibility_interpolation':'18-19 = 2/5 of HCP 15-19 band; uniform age-within-band assumption'}

DEMO_COMBOS=list(itertools.product(AGE,SEX,URBAN,EDU,ACT))

def make_support(pol,seed,attempt):
    positive=[p for p,v in pol.items() if v>1e-15]; k=len(positive)
    base=TARGET_N//k; rem=TARGET_N%k
    ns={p:base+(i<rem) for i,p in enumerate(sorted(positive,key=lambda x:(-pol[x],x)))}
    rng=np.random.default_rng(seed+attempt*100003)
    perm=np.arange(len(DEMO_COMBOS)); rng.shuffle(perm)
    rows=[]
    for ip,p in enumerate(positive):
        n=ns[p]; offset=(ip*37+attempt*11)%len(perm)
        idx=[perm[(offset+j)%len(perm)] for j in range(n)]
        if len(set(idx))<n: raise RuntimeError('support duplicate demo rows')
        for z in idx:
            a,s,u,e,act=DEMO_COMBOS[int(z)]; rows.append({'age_band':a,'sex':s,'urban_rural':u,'education_band':e,'activity_status':act,'prior_vote_or_abstention':p})
    if len(rows)!=TARGET_N: raise RuntimeError('support cardinality')
    return rows

def rake(rows,targets,max_iter=4000,tol=2e-12):
    dims=['age_band','sex','urban_rural','education_band','activity_status','prior_vote_or_abstention']
    w=np.ones(len(rows),dtype=float)/len(rows)
    for _ in range(max_iter):
        for d in dims:
            vals=targets[d]
            for cat,t in vals.items():
                idx=np.array([r[d]==cat for r in rows])
                cur=float(w[idx].sum())
                if t<=1e-18:
                    if cur: w[idx]=0.0
                else:
                    if cur<=0: return None,None
                    w[idx]*=float(t)/cur
        s=w.sum()
        if s<=0:return None,None
        w/=s
        err=max(abs(float(w[np.array([r[d]==cat for r in rows])].sum())-float(t)) for d in dims for cat,t in targets[d].items())
        if err<tol: return w,err
    return w,err

def independence_diag(rows,w,targets):
    mx=0.0
    for d in ['age_band','sex','urban_rural','education_band','activity_status']:
        for cat,dm in targets[d].items():
            for p,pm in targets['prior_vote_or_abstention'].items():
                joint=float(sum(w[i] for i,r in enumerate(rows) if r[d]==cat and r['prior_vote_or_abstention']==p))
                mx=max(mx,abs(joint-float(dm)*float(pm)))
    return mx

def choose_population(demo,pol,seed):
    targets={**demo,'prior_vote_or_abstention':pol}; best=None
    for attempt in range(300):
        rows=make_support(pol,seed,attempt); w,err=rake(rows,targets)
        if w is None: continue
        ess=float(1.0/np.sum(w*w)); mw=float(w.max()); dep=independence_diag(rows,w,targets)
        if err<=5e-10 and ess>=64 and mw<=.10:
            cand=(dep,-ess,mw,attempt,rows,w,err)
            if best is None or cand[:4]<best[:4]: best=cand
    if best is None: raise RuntimeError('no 128-archetype support satisfies frozen ESS/weight/margin gates')
    dep,negess,mw,attempt,rows,w,err=best
    for i,(r,x) in enumerate(zip(rows,w),start=1): r['archetype_id']=f'A{i:03d}'; r['weight']=float(x)
    return rows,{'raking_max_abs_error':float(err),'effective_archetype_count':float(-negess),'max_single_archetype_weight':mw,'max_demo_prior_joint_deviation_from_independence':float(dep),'support_attempt':attempt}

def margin_error(rows,targets):
    out={}
    for d,vals in targets.items():
        out[d]={}
        for cat,t in vals.items():
            obs=sum(float(r['weight']) for r in rows if r[d]==cat); out[d][cat]=float(obs-float(t))
    return out

def main():
    protocol=read(A/'protocol_v1.json')
    if protocol.get('status')!='PROTOCOL_FROZEN_BEFORE_ASV2_EXECUTION': raise RuntimeError('ASV2 protocol not frozen')
    geo=read(GEO); grows=geo['local']['rows']; split_admin=defaultdict(int)
    for r in grows: split_admin[province_key(r['prefecture_or_province'])]+=1
    b=fetch_hcp(); tabs=hcp_tables(b)
    allcert=[]
    for target,prior in [(2016,2011),(2021,2016)]:
        hrows=load_hist(prior); hmap,haudit=match_geometry_to_hist(grows,hrows)
        territories=[]; failures=[]
        for idx,g in enumerate(grows):
            cid=g['constituency_id']; prov=g['prefecture_or_province']; resolution='DIRECT_ADMIN_UNIT' if split_admin[province_key(prov)]==1 else 'PREFECTURE_OR_PROVINCE_FALLBACK_SPLIT_CONSTITUENCY'
            try:
                demo,da=demo_margin(prov,tabs); pol=political_margin(hmap[cid]); seed=260818000+target*1000+idx
                rows,qa=choose_population(demo,pol,seed)
                targets={**demo,'prior_vote_or_abstention':pol}; errs=margin_error(rows,targets)
                max_demo=max(abs(v) for d in ['age_band','sex','urban_rural','education_band','activity_status'] for v in errs[d].values())
                max_pol=max(abs(v) for v in errs['prior_vote_or_abstention'].values())
                if max_demo>.005 or max_pol>1e-9 or qa['effective_archetype_count']<64 or qa['max_single_archetype_weight']>.10: raise RuntimeError('frozen population gate failure')
                territories.append({'constituency_id':cid,'constituency_name':g['repo_name'],'prefecture_or_province':prov,'demographic_resolution':resolution,'demographic_source_year':2014,'demographic_source':HCP_URL,'prior_election_year':prior,'prior_historical_match':haudit[cid],'demographic_audit':da,'target_marginals':targets,'quality':{**qa,'max_demographic_marginal_error':float(max_demo),'max_political_marginal_error':float(max_pol)},'archetypes':rows})
            except Exception as e: failures.append({'constituency_id':cid,'error':type(e).__name__+':'+str(e)})
        result={'schema_version':'1.0','population_id':f'M26-ASV2-SYNTHETIC-POP-{target}-V1','experiment_id':'M26-AGENT-SOCIETY-V2-001','target_election_year':target,'prior_election_year':prior,'target_outcome_used':False,'hcp_source_url':HCP_URL,'hcp_source_sha256':sha(b),'archetypes_target_per_constituency':TARGET_N,'territories':territories,'failures':failures,'status':'PASS' if len(territories)==92 and not failures else 'FAIL'}
        (OUT/f'{target}_population_v1.json').write_text(json.dumps(result,ensure_ascii=False,sort_keys=True,indent=2)+'\n',encoding='utf-8')
        allcert.append({'target_year':target,'status':result['status'],'territories':len(territories),'failures':failures,'fallback_split_constituencies':sum(t['demographic_resolution'].startswith('PREFECTURE') for t in territories),'min_ess':min((t['quality']['effective_archetype_count'] for t in territories),default=0),'max_weight':max((t['quality']['max_single_archetype_weight'] for t in territories),default=1),'max_demo_error':max((t['quality']['max_demographic_marginal_error'] for t in territories),default=1),'max_political_error':max((t['quality']['max_political_marginal_error'] for t in territories),default=1),'max_demo_prior_dependence':max((t['quality']['max_demo_prior_joint_deviation_from_independence'] for t in territories),default=1)})
    cert={'schema_version':'1.0','certificate_id':'M26-ASV2-HISTORICAL-POPULATION-CERTIFICATE-V1','experiment_id':'M26-AGENT-SOCIETY-V2-001','target_outcomes_used':False,'hcp_source_sha256':sha(b),'hcp_source_url':HCP_URL,'age_eligibility_assumption':'Uniform within HCP 15-19 band; exactly 2/5 assigned to ages 18-19. This is a pre-outcome deterministic approximation and must be sensitivity-tested.','split_constituency_policy':'19 constituencies in nine split province/prefecture units inherit the official whole-unit HCP demographic marginals as an explicitly flagged hierarchical fallback; never represented as exact constituency demographics.','national_demographic_fallback_weight':0.0,'elections':allcert,'status':'PASS' if all(x['status']=='PASS' for x in allcert) else 'FAIL'}
    CERT.write_text(json.dumps(cert,ensure_ascii=False,sort_keys=True,indent=2)+'\n',encoding='utf-8'); print(json.dumps(cert,ensure_ascii=False,indent=2))
    if cert['status']!='PASS': raise SystemExit('historical population build failed')

if __name__=='__main__': main()
