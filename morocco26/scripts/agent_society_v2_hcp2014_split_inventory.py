#!/usr/bin/env python3
from __future__ import annotations
import json, re, tempfile, unicodedata
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data'/'goal100'/'agent_society_v2'/'audits'
OUT.mkdir(parents=True,exist_ok=True)
URL='https://www.rgph2014.hcp.ma/file/190479/'
TARGETS=['Azilal','Fès','Kénitra','Khémisset','Marrakech','Rabat','Salé','Taounate','Taroudant']

def norm(s):
    x=unicodedata.normalize('NFKD',str(s or ''))
    x=''.join(c for c in x if not unicodedata.combining(c)).lower()
    return re.sub(r'[^a-z0-9]+',' ',x).strip()

def main():
    r=requests.get(URL,timeout=(20,120),headers={'User-Agent':'Atlas395-ASV2/1.0'});r.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
        f.write(r.content);f.flush()
        wb=load_workbook(f.name,read_only=True,data_only=True)
        ws=wb['Indic.Ensemble']
        raw=list(ws.iter_rows(values_only=True))
    # Locate province rows by searching every cell for Province/Préfecture labels.
    targets={norm(x):x for x in TARGETS}
    sections={x:[] for x in TARGETS}; found={}
    current=None
    for ridx,row in enumerate(raw,start=1):
        texts=[str(v) for v in row[:15] if v not in (None,'')]
        ntexts=[norm(v) for v in texts]
        province_label=next((t for t in ntexts if t.startswith('province ') or t.startswith('prefecture ')),None)
        if province_label:
            match=None
            for nk,orig in targets.items():
                if nk in province_label:
                    match=orig;break
            current=match
            if match:
                found[match]={'row':ridx,'raw_first_12':[None if v is None else str(v) for v in row[:12]]}
                sections[match].append({'row':ridx,'type':'ADMIN_HEADER','raw_first_12':[None if v is None else str(v) for v in row[:12]]})
            continue
        if current:
            # stop at next region or province is handled by province_label above; if another non-target province appears current becomes None above.
            nonempty=[(i+1,str(v)) for i,v in enumerate(row[:12]) if v not in (None,'')]
            if nonempty:
                sections[current].append({'row':ridx,'type':'SUBUNIT','first_12_indexed':nonempty})
    # Keep only plausible administrative rows before metric-only corruption: subdivision name appears as text among first 12.
    for k,items in sections.items():
        filt=[]
        for item in items:
            if item['type']=='ADMIN_HEADER': filt.append(item); continue
            vals=item['first_12_indexed']
            if any(any(ch.isalpha() for ch in v) for _,v in vals): filt.append(item)
        sections[k]=filt
    result={'schema_version':'1.0','audit_id':'M26-ASV2-HCP2014-SPLIT-ADMIN-INVENTORY-V1','source_url':URL,'target_outcome_used':False,'targets':TARGETS,'found':found,'sections':sections}
    (OUT/'hcp2014_split_admin_inventory_v1.json').write_text(json.dumps(result,ensure_ascii=False,sort_keys=True,indent=2)+'\n',encoding='utf-8')
    for k in TARGETS:
        print('\n###',k,'rows',len(sections[k]))
        for x in sections[k][:120]: print(x)

if __name__=='__main__': main()
