#!/usr/bin/env python3
"""Source-level diagnostics for 2011 raw party universe and 2021 registered voters.

This script is diagnostic only. It never changes forecasts or canonical outcomes.
"""
from __future__ import annotations

import hashlib
import json
import re
import unicodedata
import urllib.parse
import urllib.request
from html.parser import HTMLParser
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
HIST = ROOT / "data" / "goal100" / "historical"
FP = ROOT / "data" / "goal100" / "forecast_pipeline"
RAW_2011 = HIST / "raw" / "parlement-elections-2011-1-0.xlsx"
CANON_2011 = HIST / "tafra_legislative_2011_canonical.json"
OUT_2011 = FP / "diagnostic_2011_raw_party_universe_v1.json"
OUT_2021 = FP / "registered_voters_2021_source_probe_v1.json"

TARGETS = {
    "PLJS": ["pljs", "parti de la liberte et de la justice sociale", "parti de la liberté et de la justice sociale"],
    "PRE": ["pre", "parti du renouveau et de l equite", "parti du renouveau et de l'équité", "parti du renouveau et de l’equite"],
    "PUD": ["pud", "parti de l unite et de la democratie", "parti de l'unité et de la démocratie", "parti de l’unite et de la democratie"],
}

OFFICIAL_2021_SEEDS = [
    "https://www.elections.ma/",
    "https://www.elections.ma/index.aspx",
    "https://www.elections.ma/elections/legislatives/resultats.aspx?IE=1&Id=T1uzm+f7U%2FWFF+rn+x03Zg%3D%3D",
    "https://www.elections.ma/faq/legislatives/faq.aspx",
]
AR_TERMS = ("الهيئة الناخبة", "المسجلين", "المقيدين", "الناخبين المقيدين")
LATIN_TERMS = ("inscrits", "electeurs", "électeurs", "corps electoral", "corps électoral")


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def norm(s: Any) -> str:
    x = "" if s is None else str(s)
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = x.casefold()
    x = re.sub(r"[\W_]+", " ", x, flags=re.UNICODE)
    return " ".join(x.split())


class LinkParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links: list[dict[str, str]] = []
        self._href: str | None = None
        self._txt: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            self._href = dict(attrs).get("href")
            self._txt = []

    def handle_data(self, data):
        if self._href is not None:
            self._txt.append(data)

    def handle_endtag(self, tag):
        if tag.lower() == "a" and self._href is not None:
            self.links.append({"href": self._href, "text": " ".join(self._txt).strip()})
            self._href = None
            self._txt = []


def fetch(url: str) -> dict[str, Any]:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; Morocco26ForecastAudit/1.0)",
            "Accept-Language": "ar,fr;q=0.9,en;q=0.7",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            b = r.read()
            ctype = r.headers.get("content-type", "")
            enc = "utf-8"
            m = re.search(r"charset=([\w-]+)", ctype, re.I)
            if m:
                enc = m.group(1)
            try:
                text = b.decode(enc, errors="replace")
            except LookupError:
                text = b.decode("utf-8", errors="replace")
            return {
                "requested_url": url,
                "final_url": r.geturl(),
                "status": getattr(r, "status", 200),
                "content_type": ctype,
                "bytes": len(b),
                "sha256": hashlib.sha256(b).hexdigest(),
                "text": text,
            }
    except Exception as e:
        return {"requested_url": url, "error": f"{type(e).__name__}: {e}"}


def audit_2011() -> dict[str, Any]:
    canon = json.loads(CANON_2011.read_text(encoding="utf-8"))
    canonical_keys = sorted({k for row in canon["rows"] for k in (row.get("votes") or {})})
    wb = load_workbook(RAW_2011, data_only=False, read_only=False)
    sheets = []
    all_strings: list[tuple[str, str, str]] = []
    previews = {}
    for ws in wb.worksheets:
        sheets.append({
            "title": ws.title,
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "hidden_columns": [k for k, dim in ws.column_dimensions.items() if dim.hidden],
        })
        sample = []
        for row in ws.iter_rows():
            nonempty = [(c.coordinate, c.value) for c in row if c.value not in (None, "")]
            if nonempty and len(sample) < 12:
                sample.append(nonempty[:40])
            for c in row:
                if isinstance(c.value, str) and c.value.strip():
                    all_strings.append((ws.title, c.coordinate, c.value.strip()))
        previews[ws.title] = sample

    norm_cells = [(s, c, v, norm(v)) for s, c, v in all_strings]
    target_hits = {}
    for code, variants in TARGETS.items():
        nv = [norm(v) for v in variants]
        hits = []
        for sheet, cell, raw, n in norm_cells:
            if n == norm(code) or any(v and v in n for v in nv[1:]):
                hits.append({"sheet": sheet, "cell": cell, "value": raw})
        target_hits[code] = hits

    short_codes = sorted({
        raw.strip().upper()
        for _, _, raw in all_strings
        if re.fullmatch(r"[A-Za-z0-9]{2,6}", raw.strip())
    })
    target_present = {k: bool(v) for k, v in target_hits.items()}
    if not any(target_present.values()):
        status = "SOURCE_INCOMPLETE_FOR_EXACT_SEAT_REPLAY"
        interpretation = (
            "PLJS/PRE/PUD are absent from every cell of the retained raw TAFRA workbook, "
            "so their absence from the canonical file is upstream of canonical ingestion."
        )
    else:
        status = "INGESTION_OR_MAPPING_REVIEW_REQUIRED"
        interpretation = (
            "At least one missing official winner code/name is present in the raw workbook; "
            "canonical ingestion/mapping must be audited before declaring the source incomplete."
        )

    return {
        "schema_version": "1.0",
        "diagnostic_id": "M26-2011-RAW-PARTY-UNIVERSE-DIAGNOSTIC-V1",
        "raw_path": str(RAW_2011.relative_to(ROOT)),
        "raw_sha256_runtime": sha256(RAW_2011),
        "canonical_path": str(CANON_2011.relative_to(ROOT)),
        "canonical_party_keys": canonical_keys,
        "canonical_party_key_count": len(canonical_keys),
        "workbook_sheets": sheets,
        "first_nonempty_rows_preview": previews,
        "short_code_like_strings_in_workbook": short_codes,
        "missing_reference_winner_targets": list(TARGETS),
        "target_hits_anywhere_in_raw_workbook": target_hits,
        "target_present": target_present,
        "status": status,
        "interpretation": interpretation,
        "F0_modified": False,
    }


def probe_2021() -> dict[str, Any]:
    pages: dict[str, dict[str, Any]] = {}
    queue = list(OFFICIAL_2021_SEEDS)
    seen: set[str] = set()
    discovered: list[dict[str, str]] = []
    max_pages = 18

    while queue and len(seen) < max_pages:
        url = queue.pop(0)
        if url in seen:
            continue
        seen.add(url)
        rec = fetch(url)
        text = rec.pop("text", "")
        lower = text.casefold()
        rec["contains_electorate_terms"] = any(t in text for t in AR_TERMS) or any(t in lower for t in LATIN_TERMS)
        rec["digit_groups_sample"] = re.findall(r"(?<!\d)\d{2,9}(?!\d)", text)[:80]
        pages[url] = rec
        if not text:
            continue

        p = LinkParser()
        try:
            p.feed(text)
        except Exception:
            pass
        for link in p.links:
            href = link.get("href") or ""
            txt = link.get("text") or ""
            absu = urllib.parse.urljoin(rec.get("final_url", url), href)
            if urllib.parse.urlparse(absu).netloc.lower().endswith("elections.ma"):
                if (
                    "legisl" in absu.lower()
                    or any(t in txt for t in AR_TERMS)
                    or any(t in txt.casefold() for t in LATIN_TERMS)
                ):
                    discovered.append({"from": url, "url": absu, "anchor": txt})
                    if absu not in seen and absu not in queue:
                        queue.append(absu)

        hints = sorted(set(re.findall(
            r"""[\"']([^\"']*(?:elect|inscrit|ناخب|مقيد|legisl)[^\"']*)[\"']""",
            text,
            flags=re.I,
        )))
        rec["inline_endpoint_hints"] = hints[:100]

    candidate_pages = []
    for url, rec in pages.items():
        if rec.get("contains_electorate_terms"):
            candidate_pages.append({
                "url": url,
                "status": rec.get("status"),
                "sha256": rec.get("sha256"),
                "bytes": rec.get("bytes"),
                "digit_groups_sample": rec.get("digit_groups_sample", [])[:30],
                "inline_endpoint_hints": rec.get("inline_endpoint_hints", [])[:30],
            })

    return {
        "schema_version": "1.0",
        "probe_id": "M26-2021-REGISTERED-VOTERS-OFFICIAL-SOURCE-PROBE-V1",
        "authority_scope": "elections.ma only",
        "official_rule_confirmed": "2021 legislative quotient uses registered voters in the constituency divided by seats.",
        "pages_fetched": pages,
        "discovered_legislative_or_electorate_links": discovered,
        "candidate_pages_with_electorate_terms": candidate_pages,
        "district_level_registered_counts_parsed": False,
        "status": (
            "OFFICIAL_SURFACE_DISCOVERED_NEEDS_ENDPOINT_EXTRACTION"
            if candidate_pages else
            "NO_MACHINE_READABLE_DISTRICT_COUNTS_FOUND_ON_PROBED_OFFICIAL_SURFACE"
        ),
        "safety_rule": "Never infer registered voters from turnout, valid votes, or seat outcomes.",
        "F0_modified": False,
    }


def main():
    FP.mkdir(parents=True, exist_ok=True)
    a = audit_2011()
    b = probe_2021()
    OUT_2011.write_text(json.dumps(a, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    OUT_2021.write_text(json.dumps(b, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(json.dumps({
        "2011_status": a["status"],
        "2011_target_present": a["target_present"],
        "2021_status": b["status"],
        "2021_pages": len(b["pages_fetched"]),
        "2021_candidate_pages": len(b["candidate_pages_with_electorate_terms"]),
    }, sort_keys=True))


if __name__ == "__main__":
    main()
