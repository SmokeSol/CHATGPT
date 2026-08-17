#!/usr/bin/env python3
"""Inventory candidate/incumbency/switch data already present in the repository.

This is schema discovery only. A keyword hit is not evidence and does not close a
historical B2 feature gap. Any candidate file must later be mapped to a dated
pre-election cutoff and source provenance before becoming forecast-eligible.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
B2 = DATA / "goal100" / "b2" / "v1"
NOW = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
MAX_BYTES = 50 * 1024 * 1024

GROUPS = {
    "CANDIDATE_IDENTITY": [r"\bcandidat", r"\bcandidate", r"t[eê]te[_ ]de[_ ]liste", r"mandataire[_ ]de[_ ]liste", r"nom[_ ]du[_ ]candidat"],
    "LIST_RANK": [r"rang[_ ]candidat", r"candidate[_ ]rank", r"list[_ ]rank", r"ordre[_ ]liste", r"position[_ ]sur[_ ]liste"],
    "INCUMBENCY": [r"\bincumbent", r"\bd[eé]put[eé]", r"\bparlementaire", r"mandat[_ ]parlementaire", r"sortant[e]?\b"],
    "PARTY_SWITCH": [r"party[_ ]switch", r"\bd[eé]fection", r"\btranshumance", r"changement[_ ]de[_ ]parti", r"\bralliement"],
    "OFFICEHOLDER_NETWORK": [r"conseiller[_ ]communal", r"pr[eé]sident[_ ]de[_ ]commune", r"officeholder", r"mandat[_ ]local", r"[eé]lu[_ ]local"],
}
COMPILED = {group: [re.compile(pattern, re.I) for pattern in patterns] for group, patterns in GROUPS.items()}


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9éèêàâîïôûùç_ ]+", " ", str(value).lower())


def classify(values: Iterable[object]) -> dict[str, list[str]]:
    hits = {group: [] for group in GROUPS}
    for raw in values:
        text = norm(raw)
        for group, patterns in COMPILED.items():
            if any(pattern.search(text) for pattern in patterns):
                hits[group].append(str(raw))
    return {group: sorted(set(values)) for group, values in hits.items() if values}


def json_keys(value, limit: int = 20_000) -> list[str]:
    keys: list[str] = []
    stack = [value]
    seen = 0
    while stack and seen < limit:
        item = stack.pop()
        seen += 1
        if isinstance(item, dict):
            keys.extend(str(key) for key in item)
            stack.extend(item.values())
        elif isinstance(item, list):
            stack.extend(item[:1000])
    return keys


def inspect_json(path: Path) -> dict:
    value = json.loads(path.read_text(encoding="utf-8"))
    keys = json_keys(value)
    return {"kind": "JSON", "observed_schema_terms": sorted(set(keys)), "keyword_groups": classify(keys)}


def inspect_csv(path: Path) -> dict:
    with path.open(encoding="utf-8-sig", newline="", errors="replace") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        samples = []
        for index, row in enumerate(reader):
            if index >= 10:
                break
            samples.extend(row)
    return {"kind": "CSV", "observed_schema_terms": header, "keyword_groups": classify([*header, *samples])}


def inspect_xlsx(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    all_terms = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        first = next(rows, ())
        header = [str(value) for value in first if value is not None]
        sample = []
        for index, row in enumerate(rows):
            if index >= 5:
                break
            sample.extend(str(value) for value in row if value is not None)
        all_terms.extend([sheet.title, *header, *sample])
        sheets.append({"sheet": sheet.title, "header": header})
    workbook.close()
    return {"kind": "XLSX", "sheets": sheets, "observed_schema_terms": sorted(set(all_terms)), "keyword_groups": classify(all_terms)}


def main() -> None:
    if not B2.exists():
        raise SystemExit("B2_CORPUS_SCAN_FAIL: B2 scaffold missing")
    candidates = []
    extensions = {".json", ".csv", ".xlsx", ".xlsm"}
    for path in sorted(DATA.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in extensions:
            continue
        if B2 in path.parents:
            continue
        size = path.stat().st_size
        if size > MAX_BYTES:
            candidates.append({
                "path": str(path.relative_to(ROOT.parent)),
                "bytes": size,
                "status": "SKIPPED_TOO_LARGE",
                "keyword_groups": {}
            })
            continue
        try:
            if path.suffix.lower() == ".json":
                result = inspect_json(path)
            elif path.suffix.lower() == ".csv":
                result = inspect_csv(path)
            else:
                result = inspect_xlsx(path)
            groups = result.pop("keyword_groups")
            if groups:
                candidates.append({
                    "path": str(path.relative_to(ROOT.parent)),
                    "bytes": size,
                    "sha256": sha(path),
                    "status": "SCHEMA_KEYWORD_CANDIDATE_REQUIRES_REVIEW",
                    "keyword_groups": groups,
                    **result,
                })
        except Exception as exc:
            candidates.append({
                "path": str(path.relative_to(ROOT.parent)),
                "bytes": size,
                "status": "PARSE_ERROR_RECORDED",
                "error": repr(exc),
                "keyword_groups": {}
            })

    usable_candidates = [row for row in candidates if row["status"] == "SCHEMA_KEYWORD_CANDIDATE_REQUIRES_REVIEW"]
    by_group = {group: [] for group in GROUPS}
    for row in usable_candidates:
        for group in row["keyword_groups"]:
            by_group[group].append(row["path"])

    report = {
        "schema_version": "1.0",
        "audit_id": "M26-GOAL100-B2-HISTORICAL-CORPUS-INVENTORY-V1",
        "created_at": NOW,
        "gate": "PASS_INVENTORY_MATCHES_FOUND" if usable_candidates else "PASS_NO_CANDIDATE_CORPUS_FOUND",
        "scope": str(DATA.relative_to(ROOT.parent)),
        "max_file_bytes": MAX_BYTES,
        "schema_keyword_candidate_files": len(usable_candidates),
        "by_feature_group": {group: sorted(paths) for group, paths in by_group.items()},
        "files": candidates,
        "forecast_eligible_files": 0,
        "historical_gap_closed": False,
        "review_contract": [
            "confirm the matched fields describe electoral candidates or officeholders rather than generic names",
            "establish election year and pre-election cutoff availability",
            "establish source tier, URL and content hash",
            "map candidate/entity and territory IDs deterministically",
            "only then create historical atomic evidence rows"
        ],
        "scientific_rule": "A schema keyword hit is discovery evidence only. It cannot produce a coefficient or change the historical identifiability gate.",
        "next_action": "Review the highest-signal matched files; accepted files reduce the 1,932-task backfill queue, rejected false positives remain recorded."
    }
    out = B2 / "historical_candidate_corpus_inventory.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    journal_candidates = [ROOT / "FIL_D_ARIANE.md", ROOT / "FIL_ARIANE.md"]
    journal = next((path for path in journal_candidates if path.exists()), journal_candidates[0])
    marker = "B2-A007 — Inventaire du corpus candidat déjà présent"
    text = journal.read_text(encoding="utf-8")
    if marker not in text:
        text += f'''\n\n### {NOW} — {marker}\n\n- Scan déterministe des JSON/CSV/XLSX existants, sans OCR ni inférence sémantique libre.\n- Fichiers à mots-clés de schéma : `{len(usable_candidates)}`; chacun reste `REQUIRES_REVIEW`.\n- Aucun fichier n’est déclaré forecast-eligible par le scan et aucun gap historique n’est fermé automatiquement.\n- Les faux positifs, erreurs de parsing et fichiers trop volumineux restent enregistrés.\n- Prochaine action exacte : examiner les matches les plus forts, prouver leur cutoff et provenance, puis seulement réduire la file de 1 932 tâches.\n'''
        journal.write_text(text, encoding="utf-8")

    print(json.dumps({
        "gate": report["gate"],
        "candidate_files": len(usable_candidates),
        "by_feature_group_counts": {group: len(paths) for group, paths in by_group.items()},
        "forecast_eligible_files": 0,
        "historical_gap_closed": False,
        "artifact": str(out.relative_to(ROOT.parent))
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
