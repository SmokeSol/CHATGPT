#!/usr/bin/env python3
from __future__ import annotations
import json,tempfile
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data'/'goal100'/'agent_society_v2'/'audits'; OUT.mkdir(parents=True,exist_ok=True)
URL='https://www.rgph2014.hcp.ma/file/190479/'

def main():
    r=requests.get(URL,timeout=(20,120),headers={'User-Agent':'Atlas395-ASV2/1.0'}); r.raise_for_status()
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
        f.write(r.content);f.flush(); wb=load_workbook(f.name,read_only=False,data_only=True)
        out={}
        for s in ['Indic.Ensemble','Indic.Urbain','Indic.Rural']:
            ws=wb[s]
            # expand merged values down/across for first 3 header rows only
            vals=[[ws.cell(row=rr,column=c).value for c in range(1,ws.max_column+1)] for rr in range(1,4)]
            for merged in ws.merged_cells.ranges:
                if merged.min_row<=3:
                    v=ws.cell(merged.min_row,merged.min_col).value
                    for rr in range(merged.min_row,min(3,merged.max_row)+1):
                        for cc in range(merged.min_col,merged.max_col+1): vals[rr-1][cc-1]=v
            cols=[]
            for c in range(ws.max_column):
                cols.append({'col':c+1,'h1':vals[0][c],'h2':vals[1][c],'h3':vals[2][c]})
            out[s]=cols
    result={'schema_version':'1.0','source_url':URL,'target_outcome_used':False,'sheets':out}
    (OUT/'hcp2014_header_extract_v1.json').write_text(json.dumps(result,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    for s,cols in out.items():
        print('\n###',s)
        for x in cols:
            h=' | '.join(str(x[k] or '') for k in ('h1','h2','h3'))
            if any(q in h.lower() for q in ['âge','age','niveau','activité','activite','sexe','population légale','population municipale']):
                print(x['col'],h)

if __name__=='__main__': main()
