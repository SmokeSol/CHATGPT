#!/usr/bin/env python3
from __future__ import annotations
import io, json, re, unicodedata
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'data'/'goal100'/'agent_society_v2'/'hcp_casablanca_key_audit_v1.json'
URL='https://www.rgph2014.hcp.ma/file/190479/'

def norm(s):
    x=unicodedata.normalize('NFKD',str(s or '')); x=''.join(c for c in x if not unicodedata.combining(c)); return re.sub(r'\s+',' ',x.lower()).strip()

def main():
    r=requests.get(URL,timeout=(20,150),headers={'User-Agent':'Atlas395-ASV2/1.0'}); r.raise_for_status()
    wb=load_workbook(io.BytesIO(r.content),read_only=True,data_only=True)
    out={'schema_version':'1.0','source':URL,'target_outcome_used':False,'sheets':{}}
    needles=('casa','moham','anfa','fida','sebaa','hassani','chock','bernous','m sick','rachid')
    for sn in ['Indic.Ensemble','Indic.Urbain','Indic.Rural']:
        ws=wb[sn]; hits=[]
        for i,row in enumerate(ws.iter_rows(min_row=1,values_only=True),start=1):
            label=row[7] if len(row)>7 else None
            if label and any(n in norm(label) for n in needles):
                hits.append({'row':i,'label':str(label),'normalized':norm(label)})
        out['sheets'][sn]=hits
    OUT.write_text(json.dumps(out,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
    print(json.dumps(out,ensure_ascii=False,indent=2))
if __name__=='__main__': main()
