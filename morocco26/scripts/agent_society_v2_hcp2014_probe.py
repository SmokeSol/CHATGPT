#!/usr/bin/env python3
from __future__ import annotations

import hashlib, json, re, tempfile, unicodedata
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "goal100" / "agent_society_v2" / "probes" / "hcp2014"
OUT.mkdir(parents=True, exist_ok=True)

FILES = {
    "individuals": "https://www.rgph2014.hcp.ma/file/190479/",
    "households": "https://www.rgph2014.hcp.ma/file/190480/",
}

TOKENS = {
    "geography": ["region", "province", "prefecture", "commune", "municipalite", "arrondissement", "centre", "urbain", "rural", "milieu"],
    "age": ["age", "15-19", "20-24", "25-29", "30-34", "35-39", "40-44", "45-49", "50-54", "55-59", "60"],
    "sex": ["sexe", "masculin", "feminin", "homme", "femme"],
    "education": ["education", "instruction", "scolaire", "primaire", "secondaire", "superieur", "alphabet"],
    "activity": ["activite", "emploi", "actif", "occupe", "chomage", "inactif"],
}


def norm(x):
    s = unicodedata.normalize("NFKD", str(x or ""))
    s = "".join(c for c in s if not unicodedata.combining(c)).lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9%+\-/]+", " ", s)).strip()


def sha256(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def workbook_summary(name: str, url: str):
    r = requests.get(url, timeout=(20, 120), allow_redirects=True, headers={"User-Agent":"Atlas395-ASV2/1.0","Accept":"*/*"})
    r.raise_for_status()
    b = r.content
    if not b.startswith(b"PK"):
        raise RuntimeError(f"{name}: response is not an XLSX zip; content_type={r.headers.get('content-type')} bytes={len(b)}")
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(b); f.flush()
        wb = load_workbook(f.name, read_only=True, data_only=True)
        sheets=[]
        corpus=[]
        for ws in wb.worksheets:
            sample=[]
            nonempty_by_row=[]
            for ridx,row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 25), values_only=True), start=1):
                vals=[v for v in row if v not in (None,"")]
                nonempty_by_row.append((len(vals),ridx))
                if vals:
                    sample.append({"row":ridx,"values":[str(v)[:240] for v in vals[:40]]})
                    corpus.extend(norm(v) for v in vals)
            header_row=max(nonempty_by_row)[1] if nonempty_by_row else None
            sheets.append({"title":ws.title,"max_row":ws.max_row,"max_column":ws.max_column,"candidate_header_row":header_row,"sample":sample[:12]})
        text="\n".join(corpus)
        token_hits={group:sorted({t for t in toks if norm(t) in text}) for group,toks in TOKENS.items()}
        return {
            "name":name,"url":url,"final_url":r.url,"status":r.status_code,
            "content_type":r.headers.get("content-type"),"bytes":len(b),"sha256":sha256(b),
            "content_disposition":r.headers.get("content-disposition"),"sheet_count":len(wb.sheetnames),
            "sheet_names":wb.sheetnames,"sheets":sheets,"token_hits":token_hits,
        }


def main():
    results=[]
    errors=[]
    for name,url in FILES.items():
        try: results.append(workbook_summary(name,url))
        except Exception as e: errors.append({"name":name,"url":url,"error":type(e).__name__+":"+str(e)})
    out={
        "schema_version":"1.0",
        "probe_id":"M26-ASV2-HCP2014-WORKBOOK-PROBE-V1",
        "generated_at":datetime.now(timezone.utc).isoformat(),
        "source_role":"AUTHORITATIVE_DEMOGRAPHIC_INPUT_CANDIDATE",
        "target_outcome_used":False,
        "files":results,"errors":errors,
        "workbooks_valid":len(results)==2 and not errors,
        "required_theme_presence":{
            g:any(r.get("token_hits",{}).get(g) for r in results)
            for g in ("geography","age","sex","education","activity")
        },
    }
    (OUT/"probe.json").write_text(json.dumps(out,ensure_ascii=False,sort_keys=True,indent=2)+"\n",encoding="utf-8")
    print(json.dumps(out,ensure_ascii=False,indent=2))
    if not out["workbooks_valid"]: raise SystemExit("HCP workbook validation failed")

if __name__=="__main__": main()
